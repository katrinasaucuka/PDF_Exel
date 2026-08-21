"""
PDF → Excel automatizācija ar AI
==================================
Šis rīks nolasa PDF failus (jebkādas struktūras — rēķinus, līgumus,
atskaites u.c.), izmantojot AI (Claude) izvelk no tiem svarīgāko
informāciju un automātiski ieraksta datus Excel failā — katram
dokumenta tipam savā lapā (sheet).

KĀ TAS STRĀDĀ:
1. Skripts izlasa visus PDF failus norādītajā mapē.
2. Katram PDF failam izvelk teksta saturu.
3. Nosūta tekstu Claude AI, kas:
   - nosaka dokumenta tipu (piem., "Rēķins", "Līgums", "Atskaite")
   - izvelk svarīgākos datus strukturētā veidā (JSON)
4. Dati tiek pievienoti Excel failam — katram dokumenta tipam
   veidojas sava lapa (sheet), datus var salīdzināt un filtrēt.

UZSTĀDĪŠANA:
    pip install pdfplumber openpyxl anthropic

LIETOŠANA:
    1. Ievieto PDF failus mapē "pdf_faili" (vai maini PDF_MAPE zemāk).
    2. Iestati savu Anthropic API atslēgu kā vides mainīgo:
       export ANTHROPIC_API_KEY="tava-atslega"
    3. Palaid: python pdf_uz_excel.py
"""

import os
import json
import glob
from datetime import datetime

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from anthropic import Anthropic

# ---------------------------------------------------------
# IESTATĪJUMI — pielāgo pēc vajadzības
# ---------------------------------------------------------
PDF_MAPE = "pdf_faili"                  # mape, kurā atrodas PDF faili
EXCEL_FAILS = "izvilktie_dati.xlsx"     # izvades Excel fails
MODELIS = "claude-sonnet-4-6"

client = Anthropic()  # automātiski nolasa ANTHROPIC_API_KEY no vides mainīgā


def izvilkt_pdf_tekstu(pdf_cels: str) -> str:
    """Nolasa visu teksta saturu no PDF faila."""
    teksts = []
    with pdfplumber.open(pdf_cels) as pdf:
        for lapa in pdf.pages:
            lapas_teksts = lapa.extract_text()
            if lapas_teksts:
                teksts.append(lapas_teksts)
    return "\n".join(teksts)


def analizet_ar_ai(teksts: str) -> dict:
    """
    Nosūta dokumenta tekstu Claude AI un saņem atpakaļ strukturētu
    JSON ar dokumenta tipu, pamatdatiem un sarakstu ar svarīgākajiem
    punktiem/nosacījumiem, kas dokumentā minēti.
    """
    prompts = f"""Tev tiks dots teksts no PDF dokumenta. Tavs uzdevums:

1. Nosaki dokumenta tipu (piem., "Rēķins", "Līgums", "Atskaite",
   "Kvīts", "Cits").
2. Izvelc pamatdatus: datumu, summu (ja ir), iesaistītās
   puses/uzņēmumus, un īsu (līdz 15 vārdu) satura kopsavilkumu.
3. Izveido sarakstu ar 3-7 SVARĪGĀKAJIEM PUNKTIEM vai nosacījumiem,
   kas dokumentā minēti (piem., termiņi, saistības, summas,
   nosacījumi, atbildības, brīdinājumi). Katrs punkts — īss, skaidrs
   teikums (līdz 20 vārdiem). Ja dokuments ir vienkāršs un tajā nav
   vairāku atsevišķu punktu (piem., īsa kvīts), atgriez tikai 1-2
   punktus vai tukšu sarakstu.

Atbildi TIKAI ar derīgu JSON, bez papildu teksta, šādā formātā:
{{
  "dokumenta_tips": "...",
  "datums": "...",
  "summa": "...",
  "puses": "...",
  "kopsavilkums": "...",
  "svarigie_punkti": ["punkts 1", "punkts 2", "..."]
}}

Ja kāda informācija dokumentā nav atrodama, ieraksti "nav norādīts".

Dokumenta teksts:
---
{teksts[:8000]}
---"""

    atbilde = client.messages.create(
        model=MODELIS,
        max_tokens=500,
        messages=[{"role": "user", "content": prompts}],
    )

    saturs = atbilde.content[0].text.strip()
    # Notīra iespējamos markdown code-block marķierus
    saturs = saturs.replace("```json", "").replace("```", "").strip()

    try:
        return json.loads(saturs)
    except json.JSONDecodeError:
        return {
            "dokumenta_tips": "Nezināms",
            "datums": "nav norādīts",
            "summa": "nav norādīts",
            "puses": "nav norādīts",
            "kopsavilkums": "AI atbildi neizdevās apstrādāt",
            "svarigie_punkti": [],
        }


def ieraksti_excel(dati: dict, faila_nosaukums: str):
    """
    Ieraksta izvilktos datus Excel failā — katram dokumenta tipam
    savā lapā. Ja fails/lapa jau eksistē, dati tiek pievienoti klāt.
    """
    if os.path.exists(EXCEL_FAILS):
        wb = load_workbook(EXCEL_FAILS)
    else:
        wb = Workbook()
        wb.remove(wb.active)  # noņem tukšo noklusējuma lapu

    lapas_nosaukums = dati.get("dokumenta_tips", "Cits")[:31]  # Excel limits 31 rakstz.

    if lapas_nosaukums not in wb.sheetnames:
        ws = wb.create_sheet(lapas_nosaukums)
        ws.append(["Fails", "Datums", "Summa", "Puses", "Kopsavilkums",
                   "Svarīgie punkti", "Apstrādāts"])
    else:
        ws = wb[lapas_nosaukums]

    # Svarīgos punktus apvieno vienā kolonnā kā numurētu sarakstu,
    # katrs punkts savā rindā (šūnas iekšienē).
    punkti = dati.get("svarigie_punkti", []) or []
    punkti_teksts = "\n".join(f"{i}. {p}" for i, p in enumerate(punkti, start=1))

    jauna_rinda = ws.max_row + 1
    ws.append([
        faila_nosaukums,
        dati.get("datums", ""),
        dati.get("summa", ""),
        dati.get("puses", ""),
        dati.get("kopsavilkums", ""),
        punkti_teksts,
        datetime.now().strftime("%Y-%m-%d %H:%M"),
    ])

    # Ieslēdz teksta aplaušanu "Svarīgie punkti" šūnai, lai
    # daudzrindu saraksts būtu lasāms
    punkti_sunna = ws.cell(row=jauna_rinda, column=6)
    punkti_sunna.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["F"].width = 50

    wb.save(EXCEL_FAILS)


def galvena():
    if not os.path.isdir(PDF_MAPE):
        print(f"Mape '{PDF_MAPE}' neeksistē. Izveido to un ievieto tajā PDF failus.")
        return

    pdf_faili = glob.glob(os.path.join(PDF_MAPE, "*.pdf"))
    if not pdf_faili:
        print(f"Mapē '{PDF_MAPE}' nav atrasts neviens PDF fails.")
        return

    print(f"Atrasti {len(pdf_faili)} PDF faili. Sākas apstrāde...\n")

    for pdf_cels in pdf_faili:
        faila_nosaukums = os.path.basename(pdf_cels)
        print(f"Apstrādā: {faila_nosaukums}")

        teksts = izvilkt_pdf_tekstu(pdf_cels)
        if not teksts.strip():
            print("  -> PDF ir tukšs vai nesalasāms, izlaižam.\n")
            continue

        dati = analizet_ar_ai(teksts)
        ieraksti_excel(dati, faila_nosaukums)

        print(f"  -> Tips: {dati.get('dokumenta_tips')}, "
              f"Datums: {dati.get('datums')}, Summa: {dati.get('summa')}\n")

    print(f"Gatavs! Visi dati saglabāti failā: {EXCEL_FAILS}")


if __name__ == "__main__":
    galvena()
